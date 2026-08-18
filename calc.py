from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "outputs" / "living_standard_panel_2012_2026"
OUTPUT_DIR = ROOT / "outputs" / "personal_living_standard_app"
YEARS = list(range(2014, 2027))
BASE_YEAR = 2014

FEATURE_COLUMNS = {
    "עשירון כלכלי": "עשירון הכנסה קבוצה",
    "מספר נפשות": "מספר נפשות קבוצה",
    "דת": "דת קבוצה",
    "מידת דתיות": "רמת דת קבוצה",
    "גיל מפרנס ראשי": "גיל מפרנס ראשי קבוצה",
    "מרכזיות ישוב": "מרכזיות ישוב קבוצה",
}

FEATURE_OPTIONS = {
    "עשירון כלכלי": ["עשירונים 1-2", "עשירונים 3-5", "עשירונים 6-8", "עשירונים 9-10"],
    "מספר נפשות": ["1 נפש", "2 נפשות", "3-5 נפשות", "6+ נפשות"],
    "דת": ["יהודי", "נוצרי", "מוסלמי", "דרוזי", "אחר"],
    "מידת דתיות": ["חילוני", "מסורתי", "דתי", "חרדי", "אורח חיים מעורב/אחר"],
    "גיל מפרנס ראשי": ["עד 29", "30-39", "40-49", "50-59", "60-66", "67+"],
    "מרכזיות ישוב": ["מרכזי/מאוד מרכזי", "בינוני", "פריפריאלי/מאוד פריפריאלי"],
}

CONSUMPTION_TO_PRICE = {
    "מזון בלי ירקות ופירות": ("ממוצע אחוז מזון בלי ירקות ופירות", "שינוי מחיר מזון בלי ירקות ופירות"),
    "ירקות ופירות": ("ממוצע אחוז ירקות ופירות", "שינוי מחיר ירקות ופירות"),
    "דיור": ("ממוצע אחוז דיור", "שינוי מחיר דיור"),
    "תחזוקת דירה": ("ממוצע אחוז תחזוקת דירה", "שינוי מחיר תחזוקת דירה"),
    "ריהוט וציוד לבית": ("ממוצע אחוז ריהוט וציוד לבית", "שינוי מחיר ריהוט וציוד לבית"),
    "הלבשה והנעלה": ("ממוצע אחוז הלבשה והנעלה", "שינוי מחיר הלבשה והנעלה"),
    "בריאות": ("ממוצע אחוז בריאות", "שינוי מחיר בריאות"),
    "חינוך תרבות ובידור": ("ממוצע אחוז חינוך תרבות ובידור", "שינוי מחיר חינוך תרבות ובידור"),
    "תחבורה ותקשורת": ("ממוצע אחוז תחבורה ותקשורת", "שינוי מחיר תחבורה ותקשורת"),
    "שונות": ("ממוצע אחוז שונות", "שינוי מחיר שונות"),
}

NO_CHOICE = "ללא בחירה"
FEATURE_CHOICES = list(FEATURE_COLUMNS) + [NO_CHOICE]


def rtl_text(value: str) -> str:
    return str(value)


def natural_text(value: str) -> str:
    return str(value)


class DataStore:
    def __init__(self) -> None:
        self.path = self._latest_panel_path()
        self.panel = pd.read_excel(self.path, sheet_name="panel")
        self.deciles = pd.read_excel(self.path, sheet_name="decile_thresholds")
        self.general_wage = pd.read_excel(self.path, sheet_name="general_wage_index")
        self._prepare()

    def _latest_panel_path(self) -> Path:
        candidates = sorted(
            DATA_DIR.glob("פאנל_רמת_חיים_קבוצות_בסיס_2014_2026*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if not candidates:
            candidates = sorted(
                ROOT.rglob("פאנל_רמת_חיים_קבוצות_בסיס_2014_2026*.xlsx"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
        for candidate in candidates:
            try:
                sheets = pd.ExcelFile(candidate).sheet_names
            except Exception:
                continue
            if {"panel", "decile_thresholds", "general_wage_index"}.issubset(sheets):
                return candidate
        raise FileNotFoundError("לא נמצא קובץ פאנל מתאים עם גיליונות panel, decile_thresholds ו-general_wage_index")

    def _prepare(self) -> None:
        self.panel["שנה"] = pd.to_numeric(self.panel["שנה"], errors="coerce").astype("Int64")
        self.panel["סכום משקולות"] = pd.to_numeric(self.panel["סכום משקולות"], errors="coerce").fillna(0.0)
        for col in ["הכנסה כוללת נטו ממוצעת", *[v[0] for v in CONSUMPTION_TO_PRICE.values()]]:
            self.panel[col] = pd.to_numeric(self.panel[col], errors="coerce")
        for col in [v[1] for v in CONSUMPTION_TO_PRICE.values()]:
            self.panel[col] = pd.to_numeric(self.panel[col], errors="coerce").fillna(0.0)
        self.general_changes = dict(
            zip(
                self.general_wage["שנה"].astype(int),
                pd.to_numeric(self.general_wage["שינוי כללי לשנה הבאה"], errors="coerce").fillna(0.0),
            )
        )

    @staticmethod
    def standard_persons(person_count: int) -> float:
        scale = {1: 1.25, 2: 2.00, 3: 2.65, 4: 3.20, 5: 3.75, 6: 4.25, 7: 4.75, 8: 5.20, 9: 5.60}
        if person_count in scale:
            return scale[person_count]
        return 5.60 + (person_count - 9) * 0.40

    def classify_decile(self, monthly_net_income: float, person_count: int) -> tuple[int, str, float]:
        standard = self.standard_persons(person_count)
        income_per_standard = monthly_net_income / standard
        for _, row in self.deciles.sort_values("עשירון").iterrows():
            upper = row["סף עליון הכנסה חודשית נטו לנפש תקנית"]
            if pd.isna(upper) or income_per_standard <= float(upper):
                return int(row["עשירון"]), str(row["קבוצת עשירונים"]), income_per_standard
        return 10, "עשירונים 9-10", income_per_standard

    def build_persona_result(self, filters: dict[str, str]) -> dict:
        filtered = self.panel.copy()
        for feature, value in filters.items():
            filtered = filtered[filtered[FEATURE_COLUMNS[feature]] == value]
        if filtered.empty:
            raise ValueError("לא נמצאו מספיק נתונים עבור קבוצת המאפיינים שנבחרה. נסו לבחור פחות מאפיינים או מאפיין אחר.")

        annual = []
        for year in YEARS:
            year_rows = filtered[filtered["שנה"] == year]
            weight = float(year_rows["סכום משקולות"].sum())
            row = {"שנה": year, "סכום משקולות": weight, "הכנסה נצפית": math.nan}
            if weight > 0:
                for section, (cons_col, _) in CONSUMPTION_TO_PRICE.items():
                    row[f"אחוז {section}"] = self._weighted_average(year_rows, cons_col, "סכום משקולות")
                row["הכנסה נצפית"] = self._weighted_average(year_rows, "הכנסה כוללת נטו ממוצעת", "סכום משקולות")
            else:
                for section in CONSUMPTION_TO_PRICE:
                    row[f"אחוז {section}"] = math.nan
            for section, (_, price_col) in CONSUMPTION_TO_PRICE.items():
                row[f"שינוי מחיר {section}"] = float(year_rows[price_col].dropna().iloc[0]) if not year_rows.empty else 0.0
            annual.append(row)

        annual_df = pd.DataFrame(annual)
        if annual_df["הכנסה נצפית"].notna().sum() == 0:
            raise ValueError("לא נמצאה אף שנה עם משקל חיובי לקבוצה שנבחרה.")

        annual_df = self._fill_consumption_composition(annual_df)
        annual_df["הכנסה למדד"] = self._fill_income_series(annual_df["הכנסה נצפית"])
        annual_df["שינוי הכנסה/שכר"] = annual_df["הכנסה למדד"].shift(-1) / annual_df["הכנסה למדד"] - 1
        last_year = annual_df["שנה"].max()
        annual_df.loc[annual_df["שנה"] == last_year, "שינוי הכנסה/שכר"] = 0.0

        inflation_values = []
        influence = {section: 0.0 for section in CONSUMPTION_TO_PRICE}
        for _, row in annual_df.iterrows():
            personal_inflation = 0.0
            if row["שנה"] == YEARS[-1]:
                inflation_values.append(0.0)
                continue
            for section in CONSUMPTION_TO_PRICE:
                share = float(row[f"אחוז {section}"]) / 100.0
                price_change = float(row[f"שינוי מחיר {section}"])
                contribution = share * price_change
                personal_inflation += contribution
                influence[section] += contribution
            inflation_values.append(personal_inflation)
        annual_df["עליית מחירים אישית"] = inflation_values

        price_index = [100.0]
        purchasing_power = [100.0]
        for index in range(len(YEARS) - 1):
            price_factor = 1 + float(annual_df.loc[index, "עליית מחירים אישית"])
            wage_factor = 1 + float(annual_df.loc[index, "שינוי הכנסה/שכר"])
            price_index.append(price_index[-1] * price_factor)
            purchasing_power.append(purchasing_power[-1] * wage_factor / price_factor)
        annual_df["מדד מחירים אישי 2014=100"] = price_index
        annual_df["מדד כוח קניה 2014=100"] = purchasing_power

        strongest = max(influence.items(), key=lambda item: item[1])
        return {"filters": filters, "annual": annual_df, "strongest_section": strongest, "source_panel": self.path}

    @staticmethod
    def _weighted_average(df: pd.DataFrame, value_col: str, weight_col: str) -> float:
        values = pd.to_numeric(df[value_col], errors="coerce")
        weights = pd.to_numeric(df[weight_col], errors="coerce")
        valid = values.notna() & weights.notna() & (weights > 0)
        if not valid.any():
            return math.nan
        return float((values[valid] * weights[valid]).sum() / weights[valid].sum())

    def _fill_consumption_composition(self, annual_df: pd.DataFrame) -> pd.DataFrame:
        filled = annual_df.copy()
        cons_cols = [f"אחוז {section}" for section in CONSUMPTION_TO_PRICE]
        filled[cons_cols] = filled[cons_cols].ffill().bfill()
        if filled[cons_cols].isna().any().any():
            raise ValueError("לא נמצא הרכב צריכה תקף לקבוצה שנבחרה.")
        return filled

    def _fill_income_series(self, observed: pd.Series) -> pd.Series:
        years = pd.Series(YEARS, index=observed.index)
        filled = observed.astype(float).copy()
        known_positions = [int(i) for i, value in filled.items() if pd.notna(value) and value > 0]
        if not known_positions:
            raise ValueError("לא נמצאה הכנסה תקפה לקבוצה שנבחרה.")

        for start, end in zip(known_positions, known_positions[1:]):
            start_income = filled.loc[start]
            end_income = filled.loc[end]
            gap = int(years.loc[end] - years.loc[start])
            for pos in range(start + 1, end):
                step = int(years.loc[pos] - years.loc[start])
                filled.loc[pos] = start_income + (end_income - start_income) * (step / gap)

        first = known_positions[0]
        for pos in range(first - 1, -1, -1):
            change = self.general_changes.get(int(years.loc[pos]), 0.0)
            filled.loc[pos] = filled.loc[pos + 1] / (1 + change)

        last = known_positions[-1]
        for pos in range(last + 1, len(filled)):
            change = self.general_changes.get(int(years.loc[pos - 1]), 0.0)
            filled.loc[pos] = filled.loc[pos - 1] * (1 + change)
        return filled


def create_result_workbook(result: dict, out_path: Path | None = None) -> Path:
    if out_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = OUTPUT_DIR / f"מדד_כוח_קניה_פרסונה_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "תוצאות"
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    title_fill = PatternFill("solid", fgColor="DDEFE8")
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    rtl_center = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
    rtl_right = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)

    ws["A1"] = rtl_text("מדד כוח הקניה לקבוצה:")
    ws["A1"].font = Font(size=16, bold=True, color="1F4E5F")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = rtl_right
    ws.merge_cells("A1:E1")

    row_idx = 2
    for feature, value in result["filters"].items():
        ws.cell(row_idx, 1, rtl_text(feature))
        ws.cell(row_idx, 2, rtl_text(value))
        ws.cell(row_idx, 1).alignment = rtl_right
        ws.cell(row_idx, 2).alignment = rtl_right
        row_idx += 1
    if not result["filters"]:
        ws.cell(row_idx, 1, rtl_text("ללא מאפיינים"))
        row_idx += 1

    note_row = row_idx + 1
    ws.cell(note_row, 1, rtl_text("הערה"))
    ws.cell(
        note_row,
        2,
        natural_text("בשנים 2024-2026 אין בטבלה מיקרו-דאטה חדש של משקי בית; הרכב הצריכה מחושב לפי השנה האחרונה שבה היו נתונים לקבוצה שלכם, ועליית השכר לפי העלייה הכללית במשק."),
    )
    ws.cell(note_row, 1).alignment = rtl_right
    ws.cell(note_row, 2).alignment = rtl_right
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=7)

    table_start = note_row + 3
    headers = [rtl_text("שנה"), rtl_text("מדד כוח קניה 2014=100"), rtl_text("מדד מחירים אישי 2014=100")]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(table_start, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = rtl_center

    data = result["annual"]
    for offset, (_, row) in enumerate(data.iterrows(), 1):
        ws.cell(table_start + offset, 1, int(row["שנה"]))
        ws.cell(table_start + offset, 2, round(float(row["מדד כוח קניה 2014=100"]), 1))
        ws.cell(table_start + offset, 3, round(float(row["מדד מחירים אישי 2014=100"]), 1))

    for row in ws.iter_rows(min_row=table_start + 1, max_row=table_start + len(data), min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.0"

    price_title_row = table_start + len(data) + 3
    ws.cell(price_title_row, 1, rtl_text("מדד המחירים לצרכן המותאם לקבוצה שלכם"))
    ws.cell(price_title_row, 1).font = Font(size=14, bold=True, color="1F4E5F")
    ws.cell(price_title_row, 1).alignment = rtl_right
    ws.merge_cells(start_row=price_title_row, start_column=1, end_row=price_title_row, end_column=5)

    strongest_section, strongest_value = result["strongest_section"]
    impact_row = price_title_row + 2
    ws.cell(impact_row, 1, natural_text("סעיף ההוצאות שהכי השפיע עליך הוא:"))
    ws.cell(impact_row, 2, natural_text(strongest_section))
    ws.cell(impact_row, 3, round(strongest_value * 100, 1))
    ws.cell(impact_row, 4, natural_text("נקודות אחוז תרומה מצטברת בקירוב"))
    for col_idx in range(1, 5):
        ws.cell(impact_row, col_idx).alignment = rtl_right

    _add_line_chart(ws, table_start, len(data), 2, rtl_text("מדד כוח הקניה"), "H3")
    _add_line_chart(ws, table_start, len(data), 3, rtl_text("מדד המחירים האישי"), "H20")

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["H"].width = 26
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = rtl_right

    sources = wb.create_sheet("מקורות ומתודולוגיה")
    sources.sheet_view.rightToLeft = True
    source_rows = [
        ("קובץ פאנל", str(result["source_panel"])),
        ("שנת בסיס", "2014=100"),
        ("חישוב מחירים", "סכום משוקלל של שינוי מחיר בכל סעיף לפי אחוז הצריכה הממוצע של הקבוצה."),
        ("חישוב כוח קניה", "מדד קודם × (1+שינוי הכנסה/שכר) / (1+עליית מחירים אישית)."),
        ("שנים חסרות", "הרכב צריכה: מילוי קדימה/אחורה לפי התצפית הקרובה. הכנסה: אינטרפולציה לינארית בין תצפיות ידועות; בקצוות שינוי כללי במשק."),
        ("ספי עשירונים", "למ״ס לוח 2, הכנסה נטו לנפש סטנדרטית: https://www.cbs.gov.il/he/publications/DocLib/2026/1994/ta2.xlsx"),
    ]
    for r_idx, (key, value) in enumerate(source_rows, 1):
        sources.cell(r_idx, 1, rtl_text(key)).font = Font(bold=True)
        sources.cell(r_idx, 1).alignment = rtl_right
        sources.cell(r_idx, 2, natural_text(value))
        sources.cell(r_idx, 2).alignment = rtl_right
    sources.column_dimensions["A"].width = 24
    sources.column_dimensions["B"].width = 120

    wb.save(out_path)
    return out_path


def _add_line_chart(ws, table_start: int, row_count: int, value_col: int, title: str, anchor: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = rtl_text("מדד")
    chart.x_axis.title = rtl_text("שנה")
    data = Reference(ws, min_col=value_col, min_row=table_start, max_row=table_start + row_count)
    years = Reference(ws, min_col=1, min_row=table_start + 1, max_row=table_start + row_count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(years)
    chart.height = 8
    chart.width = 17
    chart.style = 13
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.position = "b"
    ws.add_chart(chart, anchor)
